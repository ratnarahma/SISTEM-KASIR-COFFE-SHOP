import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk
import os

# Harga menu
menu_prices = {
    'Espresso': 12000,
    'Americano': 15000,
    'Cappucino': 20000,
    'Mochacino': 20000,
    'Caramel Macchiato': 20000,
    'Vanilla Latte': 20000,
    'Hazelnut Latte': 20000,
    'Caffe Latte': 20000,
    'Chocolatte': 20000,
    'Matcha': 20000,
    'Taro': 20000,
    'Cookies & Cream': 20000,
    'Jasmine Tea': 12000,
    'Original Tea': 10000,
    'Lemon Tea': 13000,
    'Lychee Tea': 17000,
    'Strawberry Tea': 15000
}

# Nama file Excel
excel_file = "rekapan_pesanan.xlsx"

# Buat file Excel jika belum ada
if not os.path.exists(excel_file):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Pembelian"
        ws.append(["Nama Barista", "Nama Pembeli", "Menu", "Total Harga", "Diskon", "Harga Bayar", "Uang Pembayaran", "Uang Kembali"])
        wb.save(excel_file)
    except Exception as e:
        print(f"Error creating Excel file: {e}")

class KasirApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistem Kasir Coffee Shop")

        # Sesuaikan ukuran window dengan layar penuh
        self.root.geometry(f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}")
        self.root.state("zoomed")

        # Load background image
        self.bg_image = Image.open("bg.png")
        self.bg_image = self.bg_image.resize((self.root.winfo_screenwidth(), self.root.winfo_screenheight()))
        self.bg_photo = ImageTk.PhotoImage(self.bg_image)

        self.canvas = tk.Canvas(root, width=self.root.winfo_screenwidth(), height=self.root.winfo_screenheight())
        self.canvas.pack(fill="both", expand=True)
        self.canvas.create_image(0, 0, image=self.bg_photo, anchor="nw")

        self.frame = tk.Frame(root, padx=20, pady=20, bg="#C9BCB3")
        self.frame.place(relx=0.5, rely=0.5, anchor="center")

        self.font_default = ("Times New Roman", 14)  # Font default
        self.font_small = ("Times New Roman", 10)  # Font untuk halaman kedua (lebih kecil)
        self.barista = ""
        self.pembeli = ""
        self.menu = {}
        self.uang = 0

        self.page1()

    def clear_frame(self):
        for widget in self.frame.winfo_children():
            widget.destroy()

    def page1(self):
        self.clear_frame()
        
        # Menambahkan jarak lebih antara kolom label dan entry
        tk.Label(self.frame, text="Nama Barista", font=self.font_default, bg="#C9BCB3").grid(row=0, column=0, sticky="w", padx=(0, 20))
        self.entry_barista = tk.Entry(self.frame, font=self.font_default)
        self.entry_barista.grid(row=0, column=1, padx=(0, 20))  # Menambah jarak pada kolom entry
        
        # Membuat jarak antara Nama Pembeli dan kolom input
        tk.Label(self.frame, text="Nama Pembeli", font=self.font_default, bg="#C9BCB3").grid(row=1, column=0, sticky="w", padx=(0, 20))
        self.entry_pembeli = tk.Entry(self.frame, font=self.font_default)
        self.entry_pembeli.grid(row=1, column=1, padx=(0, 20))  # Menambah jarak pada kolom entry
        
        tk.Button(self.frame, text="Next", font=self.font_default, command=self.page2).grid(row=2, columnspan=2, pady=10)

    def page2(self):
        self.barista = self.entry_barista.get()
        self.pembeli = self.entry_pembeli.get()

        if not self.barista or not self.pembeli:
            messagebox.showerror("Error", "Nama Barista dan Nama Pembeli harus diisi.")
            return

        self.clear_frame()
        tk.Label(self.frame, text="Menu", font=self.font_small, bg="#C9BCB3").grid(row=0, column=0, sticky="w")
        self.menu_vars = {}
        self.menu_entries = {}
        self.menu_temp = {}
        self.menu_sugar = {}

        for i, (menu, price) in enumerate(menu_prices.items()):
            var = tk.IntVar()
            self.menu_vars[menu] = var
            tk.Checkbutton(self.frame, text=f"{menu} - Rp {price}", variable=var, font=self.font_small, bg="#C9BCB3").grid(row=i+1, column=0, sticky="w")
            entry = tk.Entry(self.frame, font=self.font_small)
            entry.grid(row=i+1, column=1)
            self.menu_entries[menu] = entry

            temp_var = tk.StringVar(value="Ice")
            self.menu_temp[menu] = temp_var
            tk.OptionMenu(self.frame, temp_var, "Ice", "Hot").grid(row=i+1, column=2)

            sugar_var = tk.StringVar(value="Normal")
            self.menu_sugar[menu] = sugar_var
            tk.OptionMenu(self.frame, sugar_var, "Normal", "Less Sugar").grid(row=i+1, column=3)

        tk.Label(self.frame, text="Isi jumlah pesanan menggunakan angka!", fg="red", font=self.font_small, bg="#C9BCB3").grid(row=len(menu_prices)+1, column=1, sticky="w")
        tk.Button(self.frame, text="Next", font=self.font_small, command=self.page3).grid(row=len(menu_prices)+2, columnspan=4, pady=10)
        tk.Button(self.frame, text="Back", font=self.font_small, command=self.page1).grid(row=len(menu_prices)+3, columnspan=4, pady=10)

    def page3(self):
        self.menu = {}
        for menu, var in self.menu_vars.items():
            if var.get() == 1:
                try:
                    jumlah = int(self.menu_entries[menu].get())
                    if jumlah <= 0:
                        raise ValueError
                    temp = self.menu_temp[menu].get()
                    sugar = self.menu_sugar[menu].get()
                    self.menu[menu] = (jumlah, temp, sugar)
                except ValueError:
                    messagebox.showerror("Error", f"Jumlah untuk {menu} harus berupa angka positif.")
                    return

        if not self.menu:
            messagebox.showerror("Error", "Pilih setidaknya satu menu.")
            return

        self.clear_frame()
        self.total_harga = sum(menu_prices[menu] * jumlah for menu, (jumlah, temp, sugar) in self.menu.items())
        self.diskon = 0

        if self.total_harga >= 100000:
            self.diskon = self.total_harga * 0.1

        self.harga_bayar = self.total_harga - self.diskon

        tk.Label(self.frame, text="Rincian Pesanan:", font=self.font_default, bg="#C9BCB3").grid(row=0, column=0, sticky="w")
        for i, (menu, (jumlah, temp, sugar)) in enumerate(self.menu.items()):
            tk.Label(self.frame, text=f"{menu} ({temp}, {sugar}) - {jumlah} pcs", font=self.font_default, bg="#C9BCB3").grid(row=i+1, column=0, sticky="w")

        tk.Label(self.frame, text=f"Total Harga: Rp {self.total_harga}", font=self.font_default, bg="#C9BCB3").grid(row=len(self.menu)+1, column=0, sticky="w")
        tk.Label(self.frame, text=f"Diskon: Rp {self.diskon}", font=self.font_default, bg="#C9BCB3").grid(row=len(self.menu)+2, column=0, sticky="w")
        tk.Label(self.frame, text=f"Harga Bayar: Rp {self.harga_bayar}", font=self.font_default, bg="#C9BCB3").grid(row=len(self.menu)+3, column=0, sticky="w")

        tk.Label(self.frame, text="Uang Pembayaran", font=self.font_default, bg="#C9BCB3").grid(row=len(self.menu)+4, column=0, sticky="w")
        self.entry_uang = tk.Entry(self.frame, font=self.font_default)
        self.entry_uang.grid(row=len(self.menu)+4, column=1)

        tk.Button(self.frame, text="Hitung", font=self.font_default, command=self.calculate).grid(row=len(self.menu)+5, columnspan=2, pady=10)
        tk.Button(self.frame, text="Back", font=self.font_default, command=self.page2).grid(row=len(self.menu)+6, columnspan=2, pady=10)

    def calculate(self):
        try:
            self.uang = int(self.entry_uang.get())
            if self.uang <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "Uang Pembayaran harus berupa angka positif.")
            return

        if self.uang < self.harga_bayar:
            messagebox.showerror("Error", "Uang pembayaran kurang dari harga bayar.")
            return

        uang_kembali = self.uang - self.harga_bayar
        result_text = (
            f"Nama Barista: {self.barista}\n"
            f"Nama Pembeli: {self.pembeli}\n"
            f"\nPesanan:\n" +
            "\n".join([f"{menu} ({temp}, {sugar}) - {jumlah} pcs" for menu, (jumlah, temp, sugar) in self.menu.items()]) +
            f"\nTotal Harga: Rp {self.total_harga}\n"
            f"Diskon: Rp {self.diskon}\n"
            f"Harga Bayar: Rp {self.harga_bayar}\n"
            f"Uang Kembali: Rp {uang_kembali}\n"
            "\nTerima kasih telah mengunjungi Coffee Shop Jaya\nSelamat menikmati pesanan Anda"
        )

        # Simpan data ke Excel
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            data = [self.barista, self.pembeli, str(self.menu), self.total_harga, self.diskon, self.harga_bayar, self.uang, uang_kembali]
            ws.append(data)
            wb.save(excel_file)
        except Exception as e:
            messagebox.showerror("Error", f"Error saving to Excel file: {e}")
            return

        self.show_summary(result_text)

    def show_summary(self, result_text):
        self.clear_frame()
        tk.Label(self.frame, text=result_text, font=self.font_default, bg="#C9BCB3", justify="left").pack()
        tk.Button(self.frame, text="Kembali ke Halaman Awal", font=self.font_default, command=self.page1).pack(pady=10)

if __name__ == "__main__":
    root = tk.Tk()
    app = KasirApp(root)
    root.mainloop()