import os
from openpyxl import load_workbook
from tkinter import messagebox

# Nama file Excel yang menyimpan menu dan harga
menu_excel_file = "menu.xlsx"
excel_file = "rekapan_pesanan.xlsx"

# Fungsi untuk membaca menu dan harga dari file Excel
def read_menu_from_excel():
    try:
        # Memuat workbook dan sheet dari file Excel
        wb = load_workbook(menu_excel_file)
        ws = wb.active

        # Membaca menu dan harga dari sheet Excel
        menu_prices = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            menu, price = row
            menu_prices[menu] = price

        return menu_prices
    except Exception as e:
        messagebox.showerror("Error", f"Error reading menu from Excel: {e}")
        return {}

# Buat file Excel untuk pesanan jika belum ada
def create_excel_file():
    if not os.path.exists(excel_file):
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            ws.title = "Data Pembelian"
            ws.append(["Nama Barista", "Nama Pembeli", "Menu", "Total Harga", "Diskon", "Harga Bayar", "Uang Pembayaran", "Uang Kembali"])
            wb.save(excel_file)
        except Exception as e:
            print(f"Error creating Excel file: {e}")

# Fungsi untuk menyimpan data ke dalam Excel
def save_to_excel(barista, pembeli, menu, total_harga, diskon, harga_bayar, uang_pembayaran, uang_kembali):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append([barista, pembeli, str(menu), total_harga, diskon, harga_bayar, uang_pembayaran, uang_kembali])
        wb.save(excel_file)
    except Exception as e:
        raise Exception(f"Error saving to Excel file: {e}")

# Fungsi untuk menghitung total harga dan diskon
def calculate_total(menu, menu_prices):
    total_harga = sum(menu_prices[menu] * jumlah for menu, (jumlah, _, _) in menu.items())
    diskon = 0
    if total_harga >= 100000:
        diskon = total_harga * 0.1
    harga_bayar = total_harga - diskon
    return total_harga, diskon, harga_bayar

# Fungsi untuk menghitung uang kembali
def calculate_change(uang_pembayaran, harga_bayar):
    if uang_pembayaran < harga_bayar:
        raise ValueError("Uang pembayaran kurang dari harga bayar.")
    return uang_pembayaran - harga_bayar